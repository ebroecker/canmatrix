# -*- coding: utf-8 -*-
# Copyright (c) 2013, Eduard Broecker
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without modification, are permitted provided that
# the following conditions are met:
#
#    Redistributions of source code must retain the above copyright notice, this list of conditions and the
#    following disclaimer.
#    Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the
#    following disclaimer in the documentation and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED
# WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A
# PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY
# DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO,
# PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR
# OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH
# DAMAGE.

#
# this script exports xls-files from a canmatrix-object
# xls-files are the can-matrix-definitions displayed in Excel

import logging
import typing
from builtins import *

import openpyxl
import openpyxl.utils

import canmatrix
import canmatrix.formats.xls_common
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# Font Size : 8pt * 20 = 160
# font = 'font: name Arial Narrow, height 160'
font = 'font: name Verdana, height 160'

sty_header = 0
sty_norm = 0
sty_first_frame = 0
sty_white = 0

sty_green = 0
sty_green_first_frame = 0
sty_sender = 0
sty_sender_first_frame = 0
sty_sender_green = 0
sty_sender_green_first_frame = 0


def write_ecu_matrix(ecu_list, signal, frame, worksheet, row, col, first_frame):
    # type: (typing.Sequence[str], typing.Optional[canmatrix.Signal], canmatrix.Frame, xlsxwriter.workbook.Worksheet, int, int, xlsxwriter.workbook.Format) -> int
    # first-frame - style with borders:
    if first_frame == sty_first_frame:
        norm = sty_first_frame
        sender = sty_sender_first_frame
        norm_green = sty_green_first_frame
        sender_green = sty_sender_green_first_frame
    # consecutive-frame - style without borders:
    else:
        norm = sty_norm
        sender = sty_sender
        norm_green = sty_green
        sender_green = sty_sender_green

    # iterate over ECUs:
    for ecu in ecu_list:
        # every second ECU with other style
        if col % 2 == 0:
            loc_style = norm
            loc_style_sender = sender
        # every second ECU with other style
        else:
            loc_style = norm_green
            loc_style_sender = sender_green
        # write "s" "r" "r/s" if signal is sent, received or send and received by ECU
        if signal is not None and ecu in signal.receivers and ecu in frame.transmitters:
            worksheet.cell(row=row+1, column=col+1).value = "r/s"
            worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
        elif signal is not None and ecu in signal.receivers:
            worksheet.cell(row=row+1, column=col+1).value = "r"
            worksheet.cell(row=row+1, column=col+1).style = loc_style
        elif ecu in frame.transmitters:
            worksheet.cell(row=row+1, column=col+1).value = "s"
            worksheet.cell(row=row+1, column=col+1).style = loc_style_sender
        else:
            worksheet.cell(row=row+1, column=col+1).value = ""
            worksheet.cell(row=row+1, column=col+1).style = loc_style
        col += 1
    # loop over ECUs ends here
    return col


def write_excel_line(worksheet, row, col, row_array, style):
    # type: (openpyxl.workbook.Worksheet, int, int, typing.Sequence[typing.Any], xlsxwriter.workbook.Format) -> int
    for item in row_array:
        worksheet.cell(row=row+1, column=col+1).value = item
        if style != 0:
            worksheet.cell(row=row + 1, column=col + 1).style = style
        col += 1
    return col


def dump(db, filename, **options):
    # type: (canmatrix.CanMatrix, str, **str) -> None
    motorola_bit_format = options.get("xlsMotorolaBitFormat", "msbreverse")
    values_in_seperate_lines = options.get("xlsValuesInSeperateLines", True)
    additional_signal_columns = [x for x in options.get("additionalSignalAttributes", "").split(",") if x]
    additional_frame_columns = [x for x in options.get("additionalFrameAttributes", "").split(",") if x]

    head_top = [
        'ID',
        'Frame Name',
        'Cycle Time [ms]',
        'Launch Type',
        'Launch Parameter',
        'Signal Byte No.',
        'Signal Bit No.',
        'Signal Name',
        'Signal Function',
        'Signal Length [Bit]',
        'Signal Default',
        'Signal Not Available',
        'Byteorder']
    head_tail = ['Value', 'Name / Phys. Range', 'Function / Increment Unit']

    workbook = openpyxl.Workbook()
    # ws_name = os.path.basename(filename).replace('.xlsx', '')
    # worksheet = workbook.add_worksheet('K-Matrix ' + ws_name[0:22])
    worksheet = workbook.active
    worksheet.title = 'K-Matrix '
    worksheet.sheet_properties.outlinePr.summaryBelow = False

    global sty_header
    global sty_white
    global sty_first_frame
    global sty_norm
    global sty_green
    global sty_green_first_frame
    global sty_sender
    global sty_sender_first_frame
    global sty_sender_green
    global sty_sender_green_first_frame

    sty_header = NamedStyle(name="sty_header")
    sty_header.font = Font(bold=True, size=8, name='Verdana')
    sty_header.alignment = Alignment(text_rotation=90, vertical='center', horizontal='center')

    sty_first_frame = NamedStyle(name="sty_first_frame")
    sty_first_frame.font = Font(bold=True, size=8, name='Verdana', color='ff000000')
    sty_first_frame.border = Border(top=Side(border_style='thin'))

    sty_white = NamedStyle(name="sty_white")
    sty_white.font = Font(bold=True, size=8, name='Verdana', color='00ffffff')

    sty_norm = NamedStyle(name="sty_norm")
    sty_norm.font = Font(bold=True, size=8, name='Verdana', color='ff000000')

    # ECUMatrix-Styles
    sty_green = NamedStyle(name="sty_green")
    sty_green.fill = PatternFill(patternType='solid', fgColor='CCFFCC')
    # sty_green = workbook.add_format({'pattern': 1, 'fg_color': '#CCFFCC'})

    sty_green_first_frame = NamedStyle(name="sty_green_first_frame")
    sty_green_first_frame.fill = PatternFill(patternType='solid', fgColor='CCFFCC')
    sty_green_first_frame.border = Border(top=Side(border_style='thin'))

    sty_sender = NamedStyle(name="sty_sender")
    sty_sender.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0')

    sty_sender_first_frame = NamedStyle(name="sty_sender_first_frame")
    sty_sender_first_frame.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0')
    sty_sender_first_frame.border = Border(top=Side(border_style='thin'))

    sty_sender_green = NamedStyle(name="sty_sender_green")
    sty_sender_green.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0', bgColor='CCFFCC')

    sty_sender_green_first_frame = NamedStyle(name="sty_sender_green_first_frame")
    sty_sender_green_first_frame.fill = PatternFill(patternType='lightGrid', fgColor='C0C0C0', bgColor='CCFFCC')
    sty_sender_green_first_frame.border = Border(top=Side(border_style='thin'))

    row_array = head_top
    head_start = len(row_array)

    # write ECUs in first row:
    ecu_list = [ecu.name for ecu in db.ecus]
    row_array += ecu_list

    for col in range(0, len(row_array)):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col+1)] = ColumnDimension(worksheet, customWidth=True)
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 2

    row_array += head_tail

    additional_frame_start = len(row_array)

    # set width of selected Cols
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 3.57
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 21
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 12.29
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 21
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 30

    for additional_col in additional_frame_columns:
        row_array.append("frame." + additional_col)

    for additional_col in additional_signal_columns:
        row_array.append("signal." + additional_col)

    write_excel_line(worksheet, 0, 0, row_array, sty_header)

    if db.type == canmatrix.matrix_class.CAN:
        frame_hash = {}
        logger.debug("DEBUG: Length of db.frames is %d", len(db.frames))
        for frame in db.frames:
            if frame.is_complex_multiplexed:
                logger.error("Export complex multiplexers is not supported - frame %s might be uncomplete", frame.name)
            frame_hash[int(frame.arbitration_id.id)] = frame
    else:
        frame_hash = {a.name:a for a in db.frames}

    # set row to first Frame (row = 0 is header)
    row = 1

    # iterate over the frames
    for idx in sorted(frame_hash.keys()):

        frame = frame_hash[idx]
        frame_style = sty_first_frame

        # sort signals:
        sig_hash = {}
        for sig in frame.signals:
            if motorola_bit_format == "msb":
                sig_hash["%03d" % int(sig.get_startbit(bit_numbering=1)) + sig.name] = sig
            elif motorola_bit_format == "msbreverse":
                sig_hash["%03d" % int(sig.get_startbit()) + sig.name] = sig
            else:  # motorolaBitFormat == "lsb"
                sig_hash["%03d" % int(sig.get_startbit(bit_numbering=1, start_little=True)) + sig.name] = sig

        # set style for first line with border
        signal_style = sty_first_frame

        additional_frame_info = [frame.attribute(additional, default="") for additional in additional_frame_columns]

        row_array = []
        if len(sig_hash) == 0:
            row_array += canmatrix.formats.xls_common.get_frame_info(db, frame)
            for _ in range(5, head_start):
                row_array.append("")
            temp_col = write_excel_line(worksheet, row, 0, row_array, frame_style)
            temp_col = write_ecu_matrix(ecu_list, None, frame, worksheet, row, temp_col, frame_style)

            row_array = ["" for _ in range(temp_col, additional_frame_start)]
            row_array += additional_frame_info
            row_array += ["" for _ in additional_signal_columns]
            write_excel_line(worksheet, row, temp_col, row_array, frame_style)
            row += 1

        first_fold_row = row+1
        # iterate over signals
        for sig_idx in sorted(sig_hash.keys()):
            sig = sig_hash[sig_idx]

            # if not first Signal in Frame, set style
            if signal_style != sty_first_frame:
                signal_style = sty_norm

            # valuetable available?
            if len(sig.values) > 0 and not values_in_seperate_lines:
                value_style = signal_style
                # iterate over values in valuetable
                for val in sorted(sig.values.keys()):
                    row_array = canmatrix.formats.xls_common.get_frame_info(db, frame)
                    front_col = write_excel_line(worksheet, row, 0, row_array, frame_style)

                    if row >= first_fold_row:    
                        worksheet.row_dimensions[row+1].outline_level = 1

                    col = head_start
                    col = write_ecu_matrix(ecu_list, sig, frame, worksheet, row, col, frame_style)

                    # write Value
                    (frontRow, back_row) = canmatrix.formats.xls_common.get_signal(db, frame, sig, motorola_bit_format)
                    write_excel_line(worksheet, row, front_col, frontRow, signal_style)
                    back_row += additional_frame_info
                    for item in additional_signal_columns:
                        temp = getattr(sig, item, "")
                        back_row.append(temp)

                    write_excel_line(worksheet, row, col + 2, back_row, signal_style)
                    write_excel_line(worksheet, row, col, [val, sig.values[val]], value_style)

                    # no min/max here, because min/max has same col as values...
                    # next row
                    row += 1
                    # set style to normal - without border
                    signal_style = sty_white
                    frame_style = sty_white
                    value_style = sty_norm
                # loop over values ends here

            # no valuetable available
            else:
                row_array = canmatrix.formats.xls_common.get_frame_info(db, frame)
                front_col = write_excel_line(worksheet, row, 0, row_array, frame_style)
                if row >= first_fold_row:    
                    worksheet.row_dimensions[row+1].outline_level = 1

                col = head_start
                col = write_ecu_matrix(ecu_list, sig, frame, worksheet, row, col, frame_style)
                (frontRow, back_row) = canmatrix.formats.xls_common.get_signal(db, frame, sig, motorola_bit_format)
                write_excel_line(worksheet, row, front_col, frontRow, signal_style)

                if float(sig.min) != 0 or float(sig.max) != 1.0:
                    back_row.insert(0, str("%g..%g" % (sig.min, sig.max)))  # type: ignore
                else:
                    back_row.insert(0, "")
                back_row.insert(0, "")

                back_row += additional_frame_info
                for item in additional_signal_columns:
                    temp = getattr(sig, item, "")
                    back_row.append(temp)

                write_excel_line(worksheet, row, col, back_row, signal_style)
                if len(sig.values) > 0:
                    write_excel_line(worksheet, row, col, ["\n".join(["{}: {}".format(a,b) for (a,b) in sig.values.items()])], signal_style)
                # next row
                row += 1
                # set style to normal - without border
                signal_style = sty_white
                frame_style = sty_white
        # loop over signals ends here
    # loop over frames ends here

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = worksheet['B2']
    # save file
    workbook.save(filename=filename)


def load(file, **options):
    # type: (typing.BinaryIO, **str) -> canmatrix.CanMatrix

    all_ecu_names = []

    # else use this hack to read xlsx
    motorola_bit_format = options.get("xlsMotorolaBitFormat", "msbreverse")
    workbook = openpyxl.open(file)
    sheet = workbook._sheets[0]
    db = canmatrix.CanMatrix()
    # Defines not imported...
    db.add_frame_defines("GenMsgDelayTime", 'INT 0 65535')
    db.add_frame_defines("GenMsgCycleTimeActive", 'INT 0 65535')
    db.add_frame_defines("GenMsgNrOfRepetitions", 'INT 0 65535')
    launch_types = []  # type: typing.List[str]

    db.add_signal_defines("GenSigSNA", 'STRING')

    ecu_start = ecu_end = 0

    column_heads = [sheet.cell(1,a).value for a in range(1, sheet.max_column+1)]

    if 'Byteorder' in column_heads:
        ecu_start = column_heads.index('Byteorder') + 1
    else:
        ecu_start = column_heads.index('Signal Not Available') + 1

    ecu_end = column_heads.index('Value')

    # ECUs:
    for x in range(ecu_start, ecu_end):
        db.add_ecu(canmatrix.Ecu(column_heads[x]))
        all_ecu_names.append(column_heads[x])
    # initialize:
    frame_id = None
    signal_name = ""
    signal_length = 8
    new_frame = None  # type: typing.Optional[canmatrix.Frame]
    new_signal = None  # type: typing.Optional[canmatrix.Signal]

    def get_if_possible(my_row, my_value, default=None):
        if my_value in column_heads and my_row[column_heads.index(my_value)].value is not None:
            return my_row[column_heads.index(my_value)].value
        else:
            return default

    for row in sheet.rows:
        # ignore empty row
        if row[column_heads.index('ID')].value is None or row[column_heads.index('ID')].value == 'ID':
            continue

            # new frame detected
        if row[column_heads.index('ID')].value != frame_id:
            # new Frame
            frame_id = row[column_heads.index('ID')].value
            frame_name = row[column_heads.index('Frame Name')].value
            cycle_time = get_if_possible(row, 'Cycle Time [ms]', '0')
            launch_type = get_if_possible(row, 'Launch Type')
            dlc = 8
                    
            # launch_param = get_if_possible(row, 'Launch Parameter', '0')
            # launch_param = str(int(launch_param))

            if frame_id.endswith("xh"):
                new_frame = canmatrix.Frame(frame_name, canmatrix.ArbitrationId(int(frame_id[:-2], 16), extended=True), size=dlc)
            else:
                new_frame = canmatrix.Frame(frame_name, arbitration_id=int(frame_id[:-1], 16), size=dlc)

            for col_head in column_heads:
                if col_head.startswith("frame."):
                    command_str = col_head.replace("frame", "new_frame")
                    command_str += "=" + str(row[column_heads.index(col_head)].value)
                    exec(command_str)
                    
            db.add_frame(new_frame)

            # eval launch_type
            if launch_type is not None:
                new_frame.add_attribute("GenMsgSendType", launch_type)
                if launch_type not in launch_types:
                    launch_types.append(launch_type)

            new_frame.cycle_time = cycle_time

        # new signal detected
        if get_if_possible(row, 'Signal Name') != signal_name:
            receiver = []  # type: typing.List[str]
            start_byte = int(get_if_possible(row, 'Signal Byte No.', "0"))
            start_bit = int(get_if_possible(row, 'Signal Bit No.', "0"))
            signal_name = get_if_possible(row, 'Signal Name')
            signal_comment = get_if_possible(row, 'Signal Function')
            signal_length = int(get_if_possible(row, 'Signal Length [Bit]', 0))
            # signal_default = get_if_possible(row, 'Signal Default')
            # signal_sna = get_if_possible(row, 'Signal Not Available')
            multiplex = None  # type: typing.Union[str, int, None]
            if signal_comment is not None and signal_comment.startswith('Mode Signal:'):
                multiplex = 'Multiplexor'
                signal_comment = signal_comment[12:]
            elif signal_comment is not None and signal_comment.startswith('Mode '):
                mux, signal_comment = signal_comment[4:].split(':', 1)
                multiplex = int(mux.strip())

            signal_byte_order = get_if_possible(row, 'Byteorder')
            if signal_byte_order is not None:
                if 'i' in signal_byte_order:
                    is_little_endian = True
                else:
                    is_little_endian = False
            else:
                is_little_endian = True  # Default Intel

            is_signed = False

            if signal_name != "-":
                for ecu_name in all_ecu_names:
                    ecu_sender_receiver = get_if_possible(row, ecu_name)
                    if ecu_sender_receiver is not None:
                        if 's' in ecu_sender_receiver:
                            new_frame.add_transmitter(ecu_name)
                        if 'r' in ecu_sender_receiver:
                            receiver.append(ecu_name)
                new_signal = canmatrix.Signal(signal_name,
                                              start_bit=(start_byte - 1) * 8 + start_bit,
                                              size=signal_length,
                                              is_little_endian=is_little_endian,
                                              is_signed=is_signed,
                                              receivers=receiver,
                                              multiplex=multiplex)
                if not is_little_endian:
                    # motorola
                    if motorola_bit_format == "msb":
                        new_signal.set_startbit(
                            (start_byte - 1) * 8 + start_bit, bitNumbering=1)
                    elif motorola_bit_format == "msbreverse":
                        new_signal.set_startbit((start_byte - 1) * 8 + start_bit)
                    else:  # motorola_bit_format == "lsb"
                        new_signal.set_startbit(
                            (start_byte - 1) * 8 + start_bit,
                            bitNumbering=1,
                            startLittle=True
                        )
                                    
                if signal_name is not None:
                    new_frame.add_signal(new_signal)
                    new_signal.add_comment(signal_comment)
                # function = get_if_possible(row, 'Function / Increment Unit')
        value = get_if_possible(row, 'Value')
        if value is not None:
            value = str(value)
        value_name = get_if_possible(row, 'Name / Phys. Range')

        if value_name == 0 or value_name is None:
            value_name = "0"
        elif value_name == 1:
            value_name = "1"
        test = value_name
        # .encode('utf-8')

        factor = get_if_possible(row, 'Function / Increment Unit')
        if factor is not None:
            factor = factor.strip()
            if " " in factor and factor[0].isdigit():
                (factor, unit) = factor.split(" ", 1)
                factor = factor.strip()
                unit = unit.strip()
                new_signal.unit = unit
                new_signal.factor = float(factor)
            else:
                unit = factor.strip()
                new_signal.unit = unit
                new_signal.factor = 1

        if ".." in test:
            (mini, maxi) = test.strip().split("..", 2)
            try:
                new_signal.offset = new_signal.float_factory(mini)
                new_signal.min = new_signal.float_factory(mini)
                new_signal.max = new_signal.float_factory(maxi)
            except ValueError:
                new_signal.offset = 0
                new_signal.min = None
                new_signal.max = None

        elif len(value_name) > 0:
            if value is not None and value.strip():
                value_int = int(float(value))
                new_signal.add_values(value_int, value_name)
            maxi = pow(2, signal_length) - 1
            new_signal.max = float(maxi)
        else:
            new_signal.offset = 0
            new_signal.min = None
            new_signal.max = None

        for col_head in column_heads: # todo explain this possibly dangerous code with eval
            if col_head.startswith("signal."):
                command_str = col_head.replace("signal", "new_signal")
                command_str += "=" + str(row[column_heads.index(col_head)].value)
                exec(command_str)



    # dlc-estimation / dlc is not in xls, thus calculate a minimum-dlc:
    for frame in db.frames:
        frame.update_receiver()
        frame.calc_dlc()

    launch_type_enum = "ENUM"
    for launch_type in launch_types:
        if len(launch_type) > 0:
            launch_type_enum += ' "' + launch_type + '",'
    db.add_frame_defines("GenMsgSendType", launch_type_enum[:-1])

    db.set_fd_type()
    return db
